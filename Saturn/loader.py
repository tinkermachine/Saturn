import logging
logger = logging.getLogger(__name__)
from openpyxl import load_workbook, workbook
from ast_ import ASTNode, OperatorNode, RangeNode, OperandNode, FunctionNode
from cell import Cell
from tqdm import tqdm

OP_MAP = {
    '+':'+',
    '-':'-',
    '*':'*'
}

class Loader:
    """
    Class responsible for injecting an xlsx file and converting the file into Cell objects.
    1. Injects an xlsx file
    2. Uses openpyxl to convert extract formulas and value from each cell
    3. Instantiates Cell objects for each cell extracted
    """

    def __init__(self, file):
        '''
        Initializes a loader object and sets the self.file param to injected file
        @param file: xlsx file
        '''
        self.file = file
        logging.info("Loading excel file...")

        #Load file in read-only mode for faster execution. Need to read twice to extract formulas separately (Is there a better way?)
        self.wb_data_only = load_workbook(filename=self.file, data_only=True, read_only=True)
        logging.info("Values Loaded...")

        self.wb_formulas = load_workbook(filename=self.file, data_only=False, read_only=True)
        logging.info("Formulas Loaded...")

        logging.info("Excel file loaded")

        self.cellmap={}

    def getCells(self):
        '''
        Extraction of values and formulas from each source cell
        @return: Generates a val_dict for values extracted and form_dict for formulas extracted
        '''
        val_dict={}
        form_dict = {}

        logging.info("Extracting values from cells")
        for sheet in self.wb_data_only.sheetnames:
            for row in self.wb_data_only[sheet].iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # logging.debug(cell.coordinate)
                        address = "{}!{}".format(sheet,cell.coordinate)
                        val_dict[address] = cell.value

        logging.info("Extracting formulas from cells")
        for sheet in self.wb_formulas.sheetnames:
            for row in self.wb_formulas[sheet].iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # logging.debug(cell.coordinate)
                        address = "{}!{}".format(sheet, cell.coordinate)
                        form_dict[address] = cell.value

        logging.info("Retrieved list of {} values".format(len(val_dict)))
        logging.info("Retrieved list of {} formulas".format(len(form_dict)))

        #Test that values extracted match formulas extracted
        assert len(val_dict) == len (form_dict),\
            "Extraction Mismatch - Formulas extracted:{}, Valued extracted:{}".format(len(form_dict),len(val_dict))

        self.val_dict = val_dict
        self.form_dict = form_dict

    def makeCells(self):
        '''
        Wrapper function that instantiates Cell objects for each extracted cell
        @return: Returns self.cells, a list of Cell objects, created
        '''
        self.cells = []
        for (k, v), (k2, f) in zip(self.val_dict.items(), self.form_dict.items()):
            cell= self.makeCell(k)
            self.cells.append(cell)
        logging.info("{} Cell objects created".format(len(self.cells)))

        return self.cells

    def makeCell(self, address):
        '''
        Wrapper function that instantiates 1 Cell object for each extracted cell
        @return: Returns the cell object
        '''
        if address not in self.cellmap:
            cell= Cell(address)
            logging.info("-----Not in cellmap, so making cell {}------".format(address))

            cell.value = self.val_dict[address]
            cell.formula = self.form_dict[address]
            logging.info("1 Cell object created for {}".format(address))

            self.cellmap[cell.address]=cell
            return cell

        else:
            logging.info('-----Found {} in cellmap----'.format(address))
            return self.cellmap[address]



    def evaluate(self,cell):


        logging.info(">>>  Evaluating cell {}".format(cell.address))
        tree = cell.tree
        logging.info("The tree is {}".format(tree.nodes))
        print(tree.edges)

        # Check if this node is already a hardcode
        if tree.number_of_nodes() == 1:
            for root in tree:
                c = self.makeCell(root.token.value)
                ret = c.value
                logging.info("No child found. Storing value {}".format(ret))
        else:
            args =[]
            ops = []
            for node in tree:
                logging.info("****** Processing node {} *******".format(node.token.value))
                if isinstance(node, OperatorNode):
                    op = OP_MAP[node.token.value]
                    # calc_val = self.evaluateTree()

                if isinstance(node, RangeNode):
                    logging.info(">>>  Found and evaluating child {}".format(node.token.value))
                    child_cell = self.makeCell(node.token.value)
                    child_value = self.evaluate(child_cell)
                    args.append(child_value)
                    pos = tree.node[node]['pos']

            eval_str='{}{}{}'.format(args[0],op[0],args[1])
            ret = eval(eval_str)

        return ret




    # def traverseCell(self,cell):
    #     tree = cell.tree
    #
    #     for node in tree:
    #         print(node.token.type)










